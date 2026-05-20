"""Document and image understanding tool using vision-capable LLMs.

Supports:
- Image analysis via GPT-4o / Gemini Pro Vision
- Document text extraction (PDF, DOCX, TXT, CSV)
- URL-based image analysis
"""

import base64
import logging
import mimetypes
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


async def vision_analyze_handler(
    image_path: str | None = None,
    image_url: str | None = None,
    question: str = "Describe this image in detail.",
    model: str = "gpt-4o",
    api_key: str | None = None,
    **kwargs: Any,
) -> str:
    """Analyze an image using a vision-capable LLM."""
    import httpx

    if not api_key:
        import os
        api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        return "Error: No API key for vision analysis."

    # Build content array
    content: list[dict[str, Any]] = [{"type": "text", "text": question}]

    if image_path:
        path = Path(image_path)
        if not path.exists():
            return f"Error: Image not found: {image_path}"
        mime_type = mimetypes.guess_type(str(path))[0] or "image/png"
        img_bytes = path.read_bytes()
        b64 = base64.b64encode(img_bytes).decode("ascii")
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:{mime_type};base64,{b64}"},
        })
    elif image_url:
        content.append({
            "type": "image_url",
            "image_url": {"url": image_url},
        })
    else:
        return "Error: Provide either image_path or image_url."

    # Call OpenAI-compatible API
    url = "https://api.openai.com/v1/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    body = {
        "model": model,
        "messages": [{"role": "user", "content": content}],
        "max_tokens": 1024,
    }

    async with httpx.AsyncClient() as client:
        resp = await client.post(url, json=body, headers=headers, timeout=60)
        if resp.status_code != 200:
            return f"Vision API error ({resp.status_code}): {resp.text[:300]}"

        result = resp.json()
        return result.get("choices", [{}])[0].get("message", {}).get("content", "No response")


async def document_read_handler(
    file_path: str,
    workspace: str = ".",
    **kwargs: Any,
) -> str:
    """Extract text from documents (PDF, DOCX, TXT, CSV, JSON, MD)."""
    path = Path(workspace) / file_path
    if not path.exists():
        return f"Error: File not found: {file_path}"

    suffix = path.suffix.lower()

    # Plain text formats
    if suffix in (".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".log", ".xml", ".html"):
        content = path.read_text(encoding="utf-8", errors="replace")
        if len(content) > 50000:
            content = content[:50000] + "\n...[truncated]"
        return content

    # PDF
    if suffix == ".pdf":
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(path))
            text_parts = []
            for page in doc:
                text_parts.append(page.get_text())
            content = "\n\n".join(text_parts)
            if len(content) > 50000:
                content = content[:50000] + "\n...[truncated]"
            return content
        except ImportError:
            return "Error: PyMuPDF not installed. Run: pip install PyMuPDF"

    # DOCX
    if suffix == ".docx":
        try:
            import docx
            doc = docx.Document(str(path))
            paragraphs = [p.text for p in doc.paragraphs]
            content = "\n".join(paragraphs)
            if len(content) > 50000:
                content = content[:50000] + "\n...[truncated]"
            return content
        except ImportError:
            return "Error: python-docx not installed. Run: pip install python-docx"

    # XLSX
    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"--- Sheet: {sheet_name} ---")
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join(str(c) if c is not None else "" for c in row))
            content = "\n".join(lines)
            if len(content) > 50000:
                content = content[:50000] + "\n...[truncated]"
            return content
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"

    return f"Unsupported file type: {suffix}"


VISION_TOOL_SCHEMA = {
    "type": "function",
    "function": {
        "name": "analyze_image",
        "description": "Analyze an image using AI vision. Can describe images, read text from screenshots, answer questions about photos.",
        "parameters": {
            "type": "object",
            "properties": {
                "image_path": {
                    "type": "string",
                    "description": "Local file path to the image",
                },
                "image_url": {
                    "type": "string",
                    "description": "URL of the image to analyze",
                },
                "question": {
                    "type": "string",
                    "description": "Question to ask about the image",
                    "default": "Describe this image in detail.",
                },
            },
        },
    },
}

DOCUMENT_TOOL_SCHEMA = {
    "type": "function",
    "function": {
        "name": "read_document",
        "description": "Extract text from documents: PDF, DOCX, XLSX, CSV, TXT, JSON, MD, HTML.",
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the document file",
                },
            },
            "required": ["file_path"],
        },
    },
}
